from openpyxl import load_workbook
import regex

### SECTION TO BE MODIFIED BY THE USER

wb = load_workbook("grading.xlsx", data_only=True)
sheet = wb["Grading"]

course_name = "Programmation 1"
project_name = "Text Align"
project_number = "7"
file_prefix = "prg1_lab"

first_criterion_line = 4
last_criterion_line = 31
first_group_column = 4
last_group_column = 17

max_points_column = 2

### END OF SECTION

def typ_init(course, lab_nb, lab_name, student1, student2, group, grade, points, points_max):
    return f"""#import "../conf.typ": *
#let course = "{course}"
#let lab_nb = "{lab_nb}"
#let lab_name = "{lab_name}"
#let student1 = "{student1}"
#let student2 = "{student2}"
#let group = "{group}"
#let grade = "{grade}"
#let points = "{points}"
#let points_max = "{points_max}"

#show: feedback_intro.with(
    course: course,
    lab_name: lab_name, 
    lab_nb: lab_nb,
    student1: student1,
    student2: student2,
    group: group,
    grade: grade,
    points: points,
    points_max: points_max,
)
\n"""

# Returns a typst section object string
def typ_section(title, points, max_points):
    return f"""#section("{title}",{str(points)},{str(max_points)})"""

# Returns a typst subsection object string
def typ_subsection(title, points, max_points):
    return f"""#sub_section("{str(title)}",{str(points)},{str(max_points)})"""

# Returns a typst altsubsection object string
def typ_altsubsection(title, points, max_points):
    return f"""#alt_sub_section("{str(title)}")"""

# Returns a typst altsection object string
def typ_altsection(title, points, max_points):
    return f"""#alt_section("{str(title)}")"""

# Returns a typst comment object string
def typ_comment(content):
    return f"""#comment([{str(content)}])"""

# Looping over all group columns.
for group_column in range(first_group_column, last_group_column):

    # Resetting our group variables.
    total_max_points = 0
    total_points_obtained = 0
    final_grade = 0

    # Manually getting the names
    student1_name = sheet.cell(row=2,column=group_column+1).value
    student2_name = sheet.cell(row=3,column=group_column+1).value

    # Calculating the group name with the correct offset
    group_name = str(group_column-first_group_column+1)

    # Filename for feedback
    filename = f"""feedbacks/{file_prefix}{project_number}_feedback_{group_name}.typ"""

    section_func = typ_section
    subsection_func = typ_subsection

    # Opening our file to write the section, subsections and comments.
    with open(filename, "w", encoding="utf-8") as dst:

        # Iterating on every row for the current group
        for row in sheet.iter_rows(min_row=first_criterion_line, max_row=last_criterion_line, min_col=1, max_col=group_column+1):
            section = str(row[0].value)

            if section == "Pénalités diverses":
                subsection_func = typ_altsubsection
                section_func = typ_altsection

            # Checking if the row value is only Unicode letters and spaces ; If true, it's a new section.
            pattern = r'^[\p{L} ]+$'
            if section != "None" and regex.fullmatch(pattern, section):
                section_max_pts = row[max_points_column].value
                section_pts = row[group_column].value
                dst.write(section_func(section,section_pts,section_max_pts))

            # Otherwise, treat it as a subsection and take the subection name in the cell next to it.
            else:
                subsection = str(row[1].value)
                if subsection != "None":
                    section_max_pts = row[max_points_column].value
                    section_pts = row[group_column].value
                    
                    # Do not write the subtotals and grade subsections to files.
                    if subsection != "Total" and subsection != "Note" and subsection != "Note Finale":
                        dst.write(subsection_func(subsection,section_pts,section_max_pts))
                        current_comment = None
                        if row[group_column].comment:
                            current_comment = row[group_column].comment.content.split(":", 1)[1].strip()
                        if current_comment:
                            dst.write(typ_comment(current_comment))

                    # Updating our variables to write them to the file later
                    if subsection == "Total" and total_max_points == 0 and total_points_obtained == 0:
                        total_points_obtained = row[group_column].value
                        total_max_points = row[max_points_column].value
                        print("Group",group_name,"got",total_points_obtained,"points out of",total_max_points)

                    if subsection == "Note Finale":
                        final_grade = round(row[group_column].value,1)
                        print("Group",group_name,"was graded",final_grade)
    
    # Since we need to append lines before what we already wrote, read the file and copy it.
    with open(filename, "r", encoding="utf-8") as old:
        original_content = old.readlines()
    
    # Write everything in the correct order.
    with open(filename, "w", encoding="utf-8") as final:
        final.writelines(typ_init(course_name, project_number, project_name, student1_name, student2_name, group_name, final_grade, total_points_obtained, total_max_points))
        final.writelines(original_content)